import openpyxl
import numpy as np


wb = openpyxl.load_workbook('c:/Users/Administrator/Documents/code2.0/python/other/爆款数据.xlsx')
ws = wb.active

D = []
for i in range(2,ws.max_row+1):
    t = []
    for j in range(3, 7):
        t.append(ws.cell(row=i, column=j).value)
    D.append(t)
nD = np.array(D,dtype='float')
nD[:,2] = nD[:,2]*nD[:,0]       # 用价格*销量替换价格

'''
for i in range(3):              # 归一化
    nD[:,i] = (nD[:,i] - nD[:,i].min()) / (nD[:,i].max() - nD[:,i].min())
'''

# 分割训练集和测试集合
per = np.random.permutation(nD.shape[0])
test_x = nD[per[:30],:3]
test_y = nD[per[:30],3]
train_x = nD[per[30:],:3]
train_y = nD[per[30:],3]



from sklearn import svm
clf = svm.SVC(gamma='scale')
print(clf.fit(train_x, train_y))   
predict_y = clf.predict(test_x)
print('准确率',sum((predict_y==test_y)/test_y.size))


'''
## 画图
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D  # 空间三维画图
plt.rcParams['font.sans-serif'] = ['KaiTi'] # 指定默认字体
fig = plt.figure()
ax = Axes3D(fig)

ax.scatter(nD[nD[:,3] == 1,0], nD[nD[:,3] == 1,1], nD[nD[:,3] == 1,2],c='r',label='爆款')
ax.scatter(nD[nD[:,3] == 0,0], nD[nD[:,3] == 0,1], nD[nD[:,3] == 0,2],c='b',label='非爆款')

 
# 添加坐标轴
ax.set_xlabel('销量', fontdict={'size': 15, 'color': 'red'})
ax.set_ylabel('收藏', fontdict={'size': 15, 'color': 'red'})
ax.set_zlabel('价格', fontdict={'size': 15, 'color': 'red'})
plt.show()
'''